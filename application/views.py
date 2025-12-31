import openpyxl
from django.contrib.auth.decorators import login_required
from django.core.files import File
from django.core.paginator import Paginator
from django.http import Http404, HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse

from .filters import ItemFilter
from .forms import UploadFileForm
from .models import Item


def read_xlsx(file: File) -> list[tuple[int, str, str, float]]:
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    return list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))


@login_required
def index(request: HttpRequest) -> HttpResponse:
    queryset = Item.objects.annotate_search()

    filterset = ItemFilter(request.GET, queryset=queryset, request=request)
    queryset = filterset.qs

    paginator = Paginator(queryset, 10)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    template_name = "application/index.html"
    context = {
        "object_list": page_obj.object_list,
        "page_obj": page_obj,
        "filterset": filterset,
    }

    if request.htmx:
        template_name = "cotton/tbody.html"

    return render(request, template_name, context)


@login_required
def upload(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            file = form.cleaned_data["file"]

            try:
                data = read_xlsx(file)
            except Exception as e:  # noqa: BLE001
                raise Http404(str(e)) from None

            Item.objects.all().delete()

            for item in data:
                serial, code, name, initial_qty = item
                Item(serial=serial, code=code, name=name, initial_qty=initial_qty).save()

            return redirect(reverse("index"))

        return HttpResponse("ERROR")

    form = UploadFileForm()
    return render(request, "application/upload.html", {"form": form})
